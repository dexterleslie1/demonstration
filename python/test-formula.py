# 需要安装openpyxl库: pip install openpyxl
import openpyxl

excel_file = 'test-formula.xlsx'
excel_file_output = 'test-formula-output.xlsx'

# 使用data_only=True读取公式计算后的值
wb_read = openpyxl.load_workbook(excel_file, data_only=True)
ws_read = wb_read.active
cell_value = ws_read.cell(row=1, column=3).value

# 打印单元格值
print(f"第一行第三列的值是: {cell_value}")

# 使用默认模式保存文件（保留公式）
wb_save = openpyxl.load_workbook(excel_file)
wb_save.save(excel_file_output)
wb_save.close()

# 关闭读取的工作簿
wb_read.close()

# 注意：下面的代码显示为：第一行第三列的值是: None，这是因为新创建的表格没有打开过，公式没有执行并保存，所以获取到公式的结果为None
# 解决方案：1、手动打开一次Excel触发公式执行并保存。2、使用xlwings自动打开Excel触发公式执行并保存。
wb_read = openpyxl.load_workbook(excel_file_output, data_only=True)
ws_read = wb_read.active
cell_value = ws_read.cell(row=1, column=3).value
print(f"第一行第三列的值是: {cell_value}")
wb_read.close()
